import tkinter as tk
from tkinter import PhotoImage, scrolledtext
from openpyxl import load_workbook
from datetime import datetime, timedelta
import os
import socket
import subprocess
import sys

LOG_PATH = "logs/kamera_log.txt"
PORT = 80
KONTROL_SURESI = 30  # saniye
kamera_listesi = []

dark_bg = "#2e2e2e"
dark_frame_bg = "#3a3a3a"
dark_fg = "#e1e1e1"
dark_btn_bg = "#444444"
dark_btn_fg = "#ffffff"
dark_btn_active_bg = "#555555"
dark_btn_active_fg = "#ffffff"
log_bg = "#1e1e1e"
log_fg = "#cfcfcf"

def log_yaz(kamera_adi, mesaj):
    os.makedirs("logs", exist_ok=True)
    satir = f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {kamera_adi}: {mesaj}"
    with open(LOG_PATH, "a", encoding="utf-8") as f:
        f.write(satir + "\n")

def baglanti_var_mi(ip, port):
    try:
        with socket.create_connection((ip, port), timeout=2):
            return True
    except:
        return False

def kameraları_yükle():
    global kamera_listesi
    kamera_listesi.clear()
    wb = load_workbook("kameralar.xlsx")
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Sadece 3 sütun: isim, ip, grup_adi
        if row[0] is None or row[1] is None or row[2] is None:
            continue
        kamera = {
            "isim": row[0],
            "ip": row[1],
            "grup_adi": row[2],
            "gecmis": [],
            "son_durum": None,
            "son_kesinti": "Yok"
        }
        kamera_listesi.append(kamera)

def durum_renk(kamera):
    su_an = datetime.now()
    son24saat = [(z, d) for z, d in kamera["gecmis"] if z > su_an - timedelta(hours=24)]
    kamera["gecmis"] = son24saat

    if not son24saat:
        return "black"
    
    en_son_durum = son24saat[-1][1]
    kesinti_var = any(not durum for _, durum in son24saat)

    if en_son_durum:
        return "yellow" if kesinti_var else "green"
    else:
        return "black"

class Tooltip:
    def __init__(self, widget, textfunc):
        self.widget = widget
        self.textfunc = textfunc
        self.tipwindow = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        text = self.textfunc()
        if self.tipwindow or not text:
            return
        x = self.widget.winfo_rootx() + 40
        y = self.widget.winfo_rooty() + 20
        self.tipwindow = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=text, justify='left',
                         background="#ffffe0", relief='solid', borderwidth=1,
                         font=("Arial", 8))
        label.pack(ipadx=1)

    def hide_tip(self, event=None):
        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None

class KameraArayuz:
    def __init__(self, master):
        self.master = master
        master.title("Güvenlik Kamera Kontrol Programı")
        master.configure(bg=dark_bg)
        self.footer_label = tk.Label(self.master, text="© 2025 Doğancan Dönmez", font=("Segoe UI", 8), fg="#888", bg=dark_bg)
        self.footer_label.pack(side="bottom", pady=(0,5))

        self.ikonlar = {
            "green": PhotoImage(file="icons/green.png").subsample(9, 9),
            "yellow": PhotoImage(file="icons/yellow.png").subsample(9, 9),
            "black": PhotoImage(file="icons/black.png").subsample(9, 9),
        }

        ust_frame = tk.Frame(master, bg=dark_bg)
        ust_frame.pack(fill="x", padx=10, pady=(10,5))

        self.excel_ac_btn = tk.Button(
            ust_frame, text="Excel Dosyasını Aç", command=self.excel_dosyasini_ac,
            bg=dark_btn_bg, fg=dark_btn_fg,
            font=("Segoe UI", 11, "bold"),
            activebackground=dark_btn_active_bg, activeforeground=dark_btn_active_fg,
            relief="raised", bd=2
        )
        self.excel_ac_btn.pack(side="right", padx=5)

        self.yenile_btn = tk.Button(
            ust_frame, text="Yenile", command=self.guncelle,
            bg=dark_btn_bg, fg=dark_btn_fg,
            font=("Segoe UI", 11, "bold"),
            activebackground=dark_btn_active_bg, activeforeground=dark_btn_active_fg,
            relief="raised", bd=2
        )
        self.yenile_btn.pack(side="right")

        self.frame = tk.Frame(master, bg=dark_bg)
        self.frame.pack(padx=10, pady=5, fill="x")

        self.gosterimler = {}

        # Kameraları grup adına göre grupla
        gruplar = {}
        for kamera in kamera_listesi:
            grup_adi = kamera["grup_adi"]
            if grup_adi not in gruplar:
                gruplar[grup_adi] = []
            gruplar[grup_adi].append(kamera)

        # Grupları alfabetik sıraya göre göster
        for idx, grup_adi in enumerate(sorted(gruplar.keys())):
            kameralar = gruplar[grup_adi]

            kolon = tk.LabelFrame(self.frame, text=grup_adi,
                                  font=("Segoe UI", 12, "bold"),
                                  bg=dark_frame_bg, fg=dark_fg,
                                  relief="groove", bd=2, padx=10, pady=10)
            kolon.grid(row=0, column=idx, padx=15, sticky="nw")

            max_width = 14

            for i, kamera in enumerate(kameralar):
                satir = i // 3 + 1
                sutun = i % 3

                kamera_frame = tk.Frame(kolon, bg="#222222", relief="ridge", bd=1,
                                        padx=2, pady=1, width=110, height=25)
                kamera_frame.grid(row=satir, column=sutun, padx=3, pady=3, sticky="w")
                kamera_frame.grid_propagate(False)

                isim_label = tk.Label(kamera_frame, text=kamera["isim"].ljust(max_width),
                                      font=("Segoe UI", 11), anchor="w",
                                      bg="#222222", fg=dark_fg, width=max_width)
                isim_label.pack(side="left")

                ikon_label = tk.Label(kamera_frame, image=self.ikonlar["black"], bg="#222222")
                ikon_label.pack(side="right")

                def create_tooltip_func(k):
                    return lambda: (
                        f"{k['isim']}\n"
                        f"IP: {k['ip']}\n"
                        f"Durum: {'Aktif' if k['son_durum'] else 'Kapalı'}\n"
                        f"Son kesinti: {k['son_kesinti']}"
                    )

                Tooltip(kamera_frame, create_tooltip_func(kamera))

                self.gosterimler[kamera["isim"]] = ikon_label

        self.log_text = scrolledtext.ScrolledText(master, height=8, state='disabled',
                                                  font=("Courier New", 10),
                                                  bg=log_bg, fg=log_fg,
                                                  relief="sunken", bd=2)
        self.log_text.pack(fill="both", expand=True, padx=10, pady=(5,10))

        self.log_dosyasini_yukle()
        self.guncelle()

    def excel_dosyasini_ac(self):
        dosya_yolu = "kameralar.xlsx"
        if not os.path.exists(dosya_yolu):
            self.log_ekle("Hata: kameralar.xlsx dosyası bulunamadı!")
            return
        
        try:
            if sys.platform == "win32":
                os.startfile(dosya_yolu)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", dosya_yolu])
            else:
                subprocess.Popen(["xdg-open", dosya_yolu])
            self.log_ekle("Excel dosyası açıldı.")
        except Exception as e:
            self.log_ekle(f"Excel dosyası açılamadı: {e}")

    def log_dosyasini_yukle(self):
        if os.path.exists(LOG_PATH):
            with open(LOG_PATH, "r", encoding="utf-8") as f:
                onceki_loglar = f.read()
            self.log_text.configure(state='normal')
            self.log_text.insert(tk.END, onceki_loglar)
            self.log_text.configure(state='disabled')

    def log_ekle(self, mesaj):
        zaman = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        satir = f"[{zaman}] {mesaj}"
        self.log_text.configure(state='normal')
        self.log_text.insert(tk.END, satir + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state='disabled')

    def guncelle(self):
        for kamera in kamera_listesi:
            ip = kamera["ip"]
            bagli_mi = baglanti_var_mi(ip, PORT)
            kamera["gecmis"].append((datetime.now(), bagli_mi))

            onceki_durum = kamera.get("son_durum")
            kamera["son_durum"] = bagli_mi

            if onceki_durum != bagli_mi:
                if not bagli_mi:
                    kamera["son_kesinti"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    log_yaz(kamera["isim"], "Kamera bağlantısı kesildi.")
                    self.log_ekle(f"{kamera['isim']} bağlantısı kesildi.")
                else:
                    log_yaz(kamera["isim"], "Kamera bağlantısı tekrar sağlandı.")
                    self.log_ekle(f"{kamera['isim']} bağlantısı tekrar sağlandı.")
                    kamera["son_kesinti"] = "Yok"

            renk = durum_renk(kamera)
            self.gosterimler[kamera["isim"]].configure(image=self.ikonlar[renk])

        self.master.after(KONTROL_SURESI * 1000, self.guncelle)
        

if __name__ == "__main__":
    kameraları_yükle()
    root = tk.Tk()
    app = KameraArayuz(root)
    root.mainloop()
